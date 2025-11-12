import pandas as pd
import sqlite3
import re 
import os
import sys

# --- Imports da função robusta (XML) ---
import xml.etree.ElementTree as ET
try:
    from lxml import etree
    LXML_DISPONIVEL = True
except ImportError:
    LXML_DISPONIVEL = False
    print("Aviso: 'lxml' não encontrada. Usando parser XML padrão (pode ser mais lento).")

# --- Imports para Formatação Excel ---
try:
    from openpyxl import load_workbook
    from openpyxl.styles import NumberFormat
except ImportError:
    print("Aviso: 'openpyxl' não encontrado. A formatação contábil pode falhar.")

# --- Configurações (NÃO MUDA) ---
ARQUIVO_DB = 'base_compras.db'
NOME_TABELA = 'rateios_compras'
COLUNAS_CHAVE_EXCEL = {
    'codigo_fornecedor': 'Forn_Cliente',
    'numero_documento': 'Documento',
    'coluna_rateio': 'Rateio'
}
HEADER_ROW_INDEX = 1 
ACCOUNTING_FORMAT = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'


# --- Função de Limpeza de Moeda BRL ---
def to_number_brl(x):
    """Converte valores monetários BRL (com . e ,) para float."""
    try:
        if pd.isna(x): return 0.0
        s = str(x).strip().replace(' ', '') # Remove espaços
        if s == "": return 0.0
        
        if s.count(',') == 1 and s.count('.') > 0:
            s = s.replace('.', '').replace(',', '.')
        elif s.count(',') == 1:
            s = s.replace(',', '.')
        elif s.count('.') == 1 and s.count(',') == 0:
            pass 
        elif s.count('.') > 0 and s.count(',') == 0:
            s = s.replace('.', '')
        
        return float(s)
    except Exception:
        return 0.0
# --- Fim da Função ---


# --- INÍCIO DA FUNÇÃO ROBUSTA (read_spreadsheetml) ---
# (Esta função está correta)
def read_spreadsheetml(path, sheet_name, header_row=HEADER_ROW_INDEX):
    """Lê uma sheet de um XML SpreadsheetML com tentativa de recuperação e limpeza."""
    
    is_lxml = False
    if LXML_DISPONIVEL:
        try:
            parser = etree.XMLParser(recover=True, encoding='utf-8')
            tree = etree.parse(path, parser=parser)
            root = tree.getroot()
            is_lxml = True
        except Exception:
            is_lxml = False 

    if not is_lxml:
        try:
            with open(path, 'rb') as f:
                raw = f.read()
            text = None
            for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
                try:
                    text = raw.decode(enc)
                    break
                except Exception:
                    continue
            if text is None: text = raw.decode('latin-1', errors='replace')
            
            text_clean = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)
            root = ET.fromstring(text_clean)
        except Exception as e:
            raise RuntimeError(f"Falha ao parsear XML mesmo após limpeza. Erro original: {e}")

    ns_uri = None
    if root.tag.startswith('{'): 
        ns_uri = root.tag.split('}')[0].strip('{')
    
    def qname(tag):
        if is_lxml and ns_uri is None and tag in ['Workbook', 'Worksheet', 'Table', 'Row', 'Cell', 'Data']:
             return tag
        return f"{{{ns_uri}}}{tag}" if ns_uri else tag
    
    worksheets = root.findall(".//" + qname("Worksheet"))
    target_ws = None
    
    for ws in worksheets:
        if ws.attrib.get(f'{{{ns_uri}}}Name') == sheet_name or ws.attrib.get('Name') == sheet_name:
            target_ws = ws
            break

    if target_ws is None:
        raise RuntimeError(f"Planilha '{sheet_name}' não encontrada no arquivo XML.")

    table = target_ws.find(qname("Table"))
    if table is None: raise RuntimeError("Tag <Table> não encontrada dentro da Worksheet.")

    rows = []
    for row in table.findall(qname("Row")):
        cells = []
        col_index = 0
        for c in row.findall(qname("Cell")):
            idx_attr = c.attrib.get(f"{{{ns_uri}}}Index") or c.attrib.get("Index")
            if idx_attr:
                try:
                    idx = int(idx_attr) - 1
                    while col_index < idx:
                        cells.append("")
                        col_index += 1
                except Exception:
                    pass
            
            data = c.find(qname("Data"))
            val = ""
            if data is not None and data.text is not None: 
                val = data.text
            cells.append(val)
            col_index += 1
        rows.append(cells)

    if len(rows) <= header_row: 
        return pd.DataFrame() 

    maxcols = max(len(r) for r in rows)
    norm_rows = [r + [""] * (maxcols - len(r)) for r in rows]

    header = norm_rows[header_row]
    cols = [str(h).replace('\n', ' ').strip() if h is not None else "" for h in header]
    data_rows = norm_rows[header_row + 1:]
    
    df = pd.DataFrame(data_rows, columns=cols)
    return df
# --- FIM DA FUNÇÃO ROBUSTA ---


# --- Nossas Funções de Tratamento (NÃO MUDA) ---
def tratar_fornecedor(valor_coluna):
    """Quebra '000256-01-HELIBRAS' ou 'MUNIC -00-MUNICIPIO'"""
    try:
        partes = str(valor_coluna).split('-', 2)
        if len(partes) == 3:
            codigo = partes[0].strip()
            loja = partes[1].strip()
            nome = partes[2].strip()
            
            if codigo == "": 
                return pd.Series([None, loja, nome])
            return pd.Series([codigo, loja, nome])
        else:
            return pd.Series([None, None, str(valor_coluna).strip()])
    except Exception as e:
        print(f"Aviso: Falha ao tratar fornecedor '{valor_coluna}'. Erro: {e}")
        return pd.Series([None, None, str(valor_coluna).strip()])

def tratar_prf_parcela(valor_coluna):
    """
    Extrai a Parcela e o Documento da coluna 'Prf-Numero Parcela'.
    """
    try:
        s = str(valor_coluna).strip()
        match = re.search(r'([\w\d]+)\s*-\s*([\d]+)\s*-', s)
        
        if match:
            parcela = match.group(1).strip()
            documento = match.group(2).strip()
            return pd.Series([parcela, documento])
        else:
            return pd.Series([None, None])
    except:
        return pd.Series([None, None])

# --- Funções de Formatação (DATA E NÚMERO) ---
def formatar_data_br(data_str):
    if pd.isna(data_str) or data_str == '':
        return None 
    try:
        dt = pd.to_datetime(data_str)
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return data_str

def aplicar_formatacao_excel(filepath, colunas_formato):
    print(f"Aplicando formatação final (Contábil e Data) em '{filepath}'...")
    try:
        wb = load_workbook(filepath)
        ws = wb.active 
        
        col_indices = {}
        for c_idx, cell in enumerate(ws[1], 1): 
            if cell.value in colunas_formato:
                col_indices[cell.value] = c_idx
        
        if not col_indices:
            print("Aviso: Nenhuma coluna de formato encontrada no cabeçalho.")
            return

        for row in ws.iter_rows(min_row=2):
            for col_name, col_idx in col_indices.items():
                cell = row[col_idx - 1] 
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = colunas_formato[col_name]
                    except (ValueError, TypeError):
                        pass 

        wb.save(filepath)
        print("Formatação aplicada com sucesso.")
    except Exception as e:
        print(f"Aviso: Falha ao aplicar formatação OpenPyXL. O arquivo está salvo, mas não formatado. Erro: {e}")


# --- FUNÇÃO PRINCIPAL (ATUALIZADA) ---

def rodar_conciliacao(caminho_arquivo_xml, caminho_arquivo_saida):
    if not os.path.exists(caminho_arquivo_xml):
        print(f"Erro: Arquivo '{caminho_arquivo_xml}' não encontrado.")
        return False

    if not os.path.exists(ARQUIVO_DB):
        print(f"Erro: Banco de dados '{ARQUIVO_DB}' não encontrado.")
        return False

    # --- 1. Leitura e Limpeza do Arquivo A (XML) ---
    print(f"Lendo '{caminho_arquivo_xml}'...")
    try:
        print("Usando o parser robusto 'read_spreadsheetml'...")
        df_xml = read_spreadsheetml(caminho_arquivo_xml, 
                                  sheet_name="2-Titulos a pagar", 
                                  header_row=1)
        
        if df_xml.empty:
            raise Exception("O parser 'read_spreadsheetml' retornou um DataFrame vazio.")
        
        df_xml.columns = [str(c).replace('\n', ' ').strip() for c in df_xml.columns]

        df_xml['_xml_row_id'] = range(len(df_xml))

    except Exception as e:
        print(f"Erro ao ler o XML com 'read_spreadsheetml': {e}")
        return False

    print("Arquivo XML lido. Tratando colunas-chave...")
    
    df_xml[['Código', 'Loja', 'Nome do Fornecedor']] = \
        df_xml['Codigo-Nome do Fornecedor'].apply(tratar_fornecedor)
    
    df_xml[['Parcela', 'Documento']] = \
        df_xml['Prf-Numero Parcela'].apply(tratar_prf_parcela)
    
    df_xml.dropna(subset=['Documento', 'Parcela'], inplace=True)
    
    colunas_texto_xml = ['Código', 'Documento', 'Parcela', 'Loja', 'Centro Custo', 'Cta.Contabil', 'Negocio?']
    for col in colunas_texto_xml:
        if col in df_xml.columns:
            df_xml[col] = df_xml[col].astype(str).str.strip().fillna('')
    
    colunas_finais_xml = list(df_xml.columns) + ['Vlr Rateado']
    xml_keys = ['Código', 'Documento'] 

    # --- 2. Leitura do Arquivo B (SQLite) ---
    print(f"Conectando ao '{ARQUIVO_DB}' e lendo a base de compras...")
    conn = sqlite3.connect(ARQUIVO_DB)
    conn.text_factory = str
    df_compras = pd.read_sql_query(f"SELECT * FROM {NOME_TABELA}", conn, dtype=str)
    conn.close()
    
    col_forn_db = COLUNAS_CHAVE_EXCEL['codigo_fornecedor']
    col_doc_db = COLUNAS_CHAVE_EXCEL['numero_documento']

    df_compras[col_forn_db] = df_compras[col_forn_db].astype(str).str.lstrip('0').str.strip().fillna('')
    df_compras[col_doc_db] = df_compras[col_doc_db].astype(str).str.lstrip('0').str.strip().fillna('')
    df_xml['Código'] = df_xml['Código'].astype(str).str.lstrip('0').str.strip().fillna('')
    df_xml['Documento'] = df_xml['Documento'].astype(str).str.lstrip('0').str.strip().fillna('')
    
    colunas_texto_db = ['Centro Custo', 'C Contabil', 'Item Conta', 'Loja']
    for col in colunas_texto_db:
        if col in df_compras.columns:
            df_compras[col] = df_compras[col].astype(str).str.strip().fillna('')
    
    if 'Vlr.Total' in df_compras.columns:
        print("Limpando valores (BRL) do Banco de Dados...")
        df_compras['Vlr.Total'] = df_compras['Vlr.Total'].apply(to_number_brl)
    else:
        print("Aviso: Coluna 'Vlr.Total' não encontrada no DB.")
        df_compras['Vlr.Total'] = 0.0 # Cria coluna vazia para evitar erro
        
    # --- (LÓGICA ATUALIZADA) Pré-cálculo ANTES do merge ---
    print("Pré-calculando contagem de rateios do banco de dados...")
    db_counts = df_compras.groupby([col_forn_db, col_doc_db]).size().to_frame('db_match_count')
    
    print("Pré-calculando soma total (custo) do rateio por documento...")
    db_soma_doc = df_compras.groupby([col_forn_db, col_doc_db])['Vlr.Total'].sum().to_frame('Soma_Doc')
    
    # Junta os cálculos no df_compras
    df_compras = df_compras.merge(db_counts, left_on=[col_forn_db, col_doc_db], right_index=True, how='left')
    df_compras = df_compras.merge(db_soma_doc, left_on=[col_forn_db, col_doc_db], right_index=True, how='left')
    # --- FIM DA ATUALIZAÇÃO ---
        
    print(f"{len(df_compras)} registros lidos do banco de dados.")

    # --- 3. A Lógica Central (Merge) ---
    print("Iniciando a conciliação (merge)...")

    df_merged = pd.merge(
        df_xml,
        df_compras,
        left_on=xml_keys,
        right_on=[col_forn_db, col_doc_db],
        how='left', 
        indicator=True,
        suffixes=('_xml', '_db') 
    )
    
    df_merged['db_match_count'] = df_merged['db_match_count'].fillna(0)
    df_merged['Soma_Doc'] = df_merged['Soma_Doc'].fillna(0) # Preenche Soma_Doc para não-matches


    # --- 4. Separação (LÓGICA ATUALIZADA) ---
    print("Separando linhas com e sem rateio...")

    def renomear_colunas(df):
        cols_para_renomear = {col: col.replace('_xml', '') for col in df.columns if col.endswith('_xml')}
        df.rename(columns=cols_para_renomear, inplace=True)
        return df

    # CASO 1: "Sem Rateio" (O Documento foi encontrado 0 ou 1 vez no DB)
    condicao_sem_rateio = (df_merged['db_match_count'] <= 1)
    
    df_final_sem_rateio = df_merged[condicao_sem_rateio].copy()
    
    df_final_sem_rateio.drop_duplicates(subset=['_xml_row_id'], keep='first', inplace=True)
    
    df_final_sem_rateio = renomear_colunas(df_final_sem_rateio)
    
    df_final_sem_rateio['Vlr Rateado'] = df_final_sem_rateio['Titulos a vencer Valor nominal'].apply(to_number_brl)
    
    df_final_sem_rateio = df_final_sem_rateio[colunas_finais_xml]

    # CASO 2: "Com Rateio" (O Documento foi encontrado 2x ou mais no BD)
    condicao_com_rateio = (df_merged['db_match_count'] > 1)
    df_final_com_rateio = df_merged[condicao_com_rateio].copy()

    # --- 5. Mapeamento e AGRUPAMENTO (LÓGICA ATUALIZADA) ---
    if not df_final_com_rateio.empty:
        print("Agrupando e somando títulos rateados...")

        # --- (ATUALIZAÇÃO) 'Soma_Doc' já foi pré-calculado. Removemos o cálculo daqui. ---
        
        col_titulos_vencer_xml = 'Titulos a vencer Valor nominal_xml' if 'Titulos a vencer Valor nominal_xml' in df_final_com_rateio.columns else 'Titulos a vencer Valor nominal'
        df_final_com_rateio['Valor_Pago_Num'] = df_final_com_rateio[col_titulos_vencer_xml].apply(to_number_brl)
        
        col_cc_db = 'Centro Custo_db' if 'Centro Custo_db' in df_final_com_rateio.columns else 'Centro Custo'
        col_cta_db = 'C Contabil' 
        col_item_db = 'Item Conta' 
        col_loja_db = 'Loja_db' if 'Loja_db' in df_final_com_rateio.columns else 'Loja'
        col_vlr_db = 'Vlr.Total' 

        grouping_keys = ['_xml_row_id', col_item_db, col_cc_db, col_cta_db, col_loja_db]
        
        agg_funcs = { 
            col_vlr_db: 'sum', 
            'Soma_Doc': 'first', # <-- 'Soma_Doc' agora vem do pré-cálculo
            'Valor_Pago_Num': 'first',
            'Código': 'first', 
            'Documento': 'first',
            'Parcela': 'first'
        }
        
        xml_cols_to_keep = [col for col in colunas_finais_xml if col not in ['Código', 'Documento', 'Parcela', 'Vlr Rateado', '_xml_row_id']]
        
        for col in df_final_com_rateio.columns:
            if col.endswith('_xml') and col not in agg_funcs:
                agg_funcs[col] = 'first'
            elif col in xml_cols_to_keep and col not in agg_funcs:
                agg_funcs[col] = 'first'

        grouping_keys_existentes = [key for key in grouping_keys if key in df_final_com_rateio.columns]
        
        print(f"Agrupando rateios por {grouping_keys_existentes}...")
        df_agrupado = df_final_com_rateio.groupby(grouping_keys_existentes, as_index=False).agg(agg_funcs)

        # 4. Calcular Proporção e 'Vlr Rateado' FINAL
        df_agrupado['Proporcao'] = 0.0
        mask_soma_valida = df_agrupado['Soma_Doc'] != 0
        df_agrupado.loc[mask_soma_valida, 'Proporcao'] = df_agrupado[col_vlr_db] / df_agrupado['Soma_Doc']
        
        df_agrupado['Vlr Rateado'] = df_agrupado['Proporcao'] * df_agrupado['Valor_Pago_Num']
        
        df_agrupado['Valor Original'] = df_agrupado[col_vlr_db]
        
        if col_cc_db in df_agrupado.columns:
            df_agrupado['Centro Custo_xml'] = df_agrupado[col_cc_db]
        if col_cta_db in df_agrupado.columns:
            col_cta_xml = 'Cta.Contabil_xml' if 'Cta.Contabil_xml' in df_agrupado.columns else 'Cta.Contabil'
            df_agrupado[col_cta_xml] = df_agrupado[col_cta_db]
        if col_item_db in df_agrupado.columns:
            col_negocio_xml = 'Negocio?_xml' if 'Negocio?_xml' in df_agrupado.columns else 'Negocio?'
            df_agrupado[col_negocio_xml] = df_agrupado[col_item_db]
        if col_loja_db in df_agrupado.columns:
            df_agrupado['Loja_xml'] = df_agrupado[col_loja_db]
        
        df_final_com_rateio = renomear_colunas(df_agrupado)
        df_final_com_rateio = df_final_com_rateio[colunas_finais_xml]
    
    else:
        print("Nenhum título com rateio (múltiplas linhas) foi encontrado na base.")

    # --- 6. Junção e Finalização ---
    print("Consolidando relatório final...")
    df_final = pd.concat([df_final_sem_rateio, df_final_com_rateio], ignore_index=True)
    
    print("Formatando colunas de data para DD/MM/YYYY...")
    colunas_data = ['Data de Emissao', 'Data de Vencto', 'Vencto Real']
    for col in colunas_data:
        if col in df_final.columns:
            df_final[col] = df_final[col].apply(formatar_data_br)
    
    colunas_novas = ['Código', 'Loja', 'Nome do Fornecedor', 'Documento', 'Parcela']
    
    colunas_originais_xml = list(df_xml.drop(columns=colunas_novas + ['_xml_row_id'], errors='ignore').columns)
    
    ordem_final = colunas_novas + colunas_originais_xml + ['Vlr Rateado']
    
    colunas_existentes_na_ordem = [col for col in ordem_final if col in df_final.columns]
    
    df_final = df_final[colunas_existentes_na_ordem]
    
    print(f"Salvando relatório em '{caminho_arquivo_saida}'...")
    df_final.to_excel(caminho_arquivo_saida, index=False)
    
    colunas_para_formatar = {
        'Valor Original': ACCOUNTING_FORMAT,
        'Tit Vencidos Valor nominal': ACCOUNTING_FORMAT,
        'Tit Vencidos Valor corrigido': ACCOUNTING_FORMAT,
        'Titulos a vencer Valor nominal': ACCOUNTING_FORMAT,
        'Vlr Rateado': ACCOUNTING_FORMAT 
    }
    
    aplicar_formatacao_excel(caminho_arquivo_saida, colunas_para_formatar)
    
    print("\n--- SUCESSO! ---")
    print("Relatório final gerado com sucesso.")
    return True