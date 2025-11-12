import pandas as pd
import re 
import os
import sys
import io 

# --- FIX: Adiciona o diretório do script ao path ---
try:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    SCRIPT_DIR = os.getcwd()
if SCRIPT_DIR not in sys.path:
    sys.path.append(SCRIPT_DIR)
# --- FIM DO FIX ---

from firebase_utils import get_db # <-- REMOVA O PONTO

# --- Imports da função robusta (XML) ---
import xml.etree.ElementTree as ET
try:
    from lxml import etree
    LXML_DISPONIVEL = True
except ImportError:
    LXML_DISPONIVEL = False
    
# --- Imports para Formatação Excel ---
try:
    from openpyxl import load_workbook
    from openpyxl.styles import NumberFormat
    OPENPYXL_DISPONIVEL = True
except ImportError:
    OPENPYXL_DISPONIVEL = False

# --- Configurações (NÃO MUDA) ---
COLECAO_FIRESTORE = 'base_compras' # Nome da coleção no Firebase
COLUNAS_CHAVE_EXCEL = {
    'codigo_fornecedor': 'Forn_Cliente',
    'numero_documento': 'Documento',
    'coluna_rateio': 'Rateio'
}
HEADER_ROW_INDEX = 1 
ACCOUNTING_FORMAT = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

# --- Funções de Limpeza e Tratamento (SEM MUDANÇAS) ---
def to_number_brl(x):
    try:
        if pd.isna(x): return 0.0
        s = str(x).strip().replace(' ', '')
        if s == "": return 0.0
        if s.count(',') == 1 and s.count('.') > 0:
            s = s.replace('.', '').replace(',', '.')
        elif s.count(',') == 1:
            s = s.replace(',', '.')
        elif s.count('.') == 1 and s.count(',') == 0:
            pass 
        elif s.count('.') > 0 and s.count(',') == 0:
            s = s.replace('.', '')
        return float(s)
    except Exception:
        return 0.0

# (Funções read_spreadsheetml, tratar_fornecedor, tratar_prf_parcela, formatar_data_br
# e aplicar_formatacao_excel permanecem IDÊNTICAS às da sua última versão.
# Apenas copie e cole elas aqui para economizar espaço.)

def read_spreadsheetml(path, sheet_name, header_row=HEADER_ROW_INDEX):
    # ... (cole sua função robusta aqui) ...
    is_lxml = False
    if LXML_DISPONIVEL:
        try:
            parser = etree.XMLParser(recover=True, encoding='utf-8')
            tree = etree.parse(path, parser=parser)
            root = tree.getroot()
            is_lxml = True
        except Exception:
            is_lxml = False 
    if not is_lxml:
        try:
            with open(path, 'rb') as f: raw = f.read()
        except Exception:
            path.seek(0); raw = path.read()
        text = None
        for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
            try: text = raw.decode(enc); break
            except Exception: continue
        if text is None: text = raw.decode('latin-1', errors='replace')
        text_clean = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)
        root = ET.fromstring(text_clean)
    ns_uri = None
    if root.tag.startswith('{'): ns_uri = root.tag.split('}')[0].strip('{')
    def qname(tag):
        if is_lxml and ns_uri is None and tag in ['Workbook', 'Worksheet', 'Table', 'Row', 'Cell', 'Data']: return tag
        return f"{{{ns_uri}}}{tag}" if ns_uri else tag
    worksheets = root.findall(".//" + qname("Worksheet"))
    target_ws = None
    for ws in worksheets:
        if ws.attrib.get(f'{{{ns_uri}}}Name') == sheet_name or ws.attrib.get('Name') == sheet_name:
            target_ws = ws; break
    if target_ws is None: raise RuntimeError(f"Planilha '{sheet_name}' não encontrada.")
    table = target_ws.find(qname("Table"))
    if table is None: raise RuntimeError("Tag <Table> não encontrada.")
    rows = []
    for row in table.findall(qname("Row")):
        cells = []; col_index = 0
        for c in row.findall(qname("Cell")):
            idx_attr = c.attrib.get(f"{{{ns_uri}}}Index") or c.attrib.get("Index")
            if idx_attr:
                try:
                    idx = int(idx_attr) - 1
                    while col_index < idx: cells.append(""); col_index += 1
                except Exception: pass
            data = c.find(qname("Data")); val = ""
            if data is not None and data.text is not None: val = data.text
            cells.append(val); col_index += 1
        rows.append(cells)
    if len(rows) <= header_row: return pd.DataFrame() 
    maxcols = max(len(r) for r in rows)
    norm_rows = [r + [""] * (maxcols - len(r)) for r in rows]
    header = norm_rows[header_row]
    cols = [str(h).replace('\n', ' ').strip() if h is not None else "" for h in header]
    data_rows = norm_rows[header_row + 1:]
    return pd.DataFrame(data_rows, columns=cols)

def tratar_fornecedor(valor_coluna):
    # ... (cole sua função aqui) ...
    try:
        partes = str(valor_coluna).split('-', 2)
        if len(partes) == 3:
            codigo = partes[0].strip()
            loja = partes[1].strip()
            nome = partes[2].strip()
            if codigo == "": return pd.Series([None, loja, nome])
            return pd.Series([codigo, loja, nome])
        else:
            return pd.Series([None, None, str(valor_coluna).strip()])
    except Exception as e:
        return pd.Series([None, None, str(valor_coluna).strip()])

def tratar_prf_parcela(valor_coluna):
    # ... (cole sua função aqui) ...
    try:
        s = str(valor_coluna).strip()
        match = re.search(r'([\w\d]+)\s*-\s*([\d]+)\s*-', s)
        if match:
            parcela = match.group(1).strip()
            documento = match.group(2).strip()
            return pd.Series([parcela, documento])
        else:
            return pd.Series([None, None])
    except:
        return pd.Series([None, None])

def formatar_data_br(data_str):
    # ... (cole sua função aqui) ...
    if pd.isna(data_str) or data_str == '': return None 
    try:
        dt = pd.to_datetime(data_str)
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return data_str

def aplicar_formatacao_excel(workbook, colunas_formato):
    # ... (cole sua função aqui) ...
    print(f"Aplicando formatação final (Contábil e Data) no Workbook...")
    if not OPENPYXL_DISPONIVEL:
        print("Aviso: OpenPyXL não disponível. Pulando formatação.")
        return workbook
    try:
        ws = workbook.active 
        col_indices = {}
        for c_idx, cell in enumerate(ws[1], 1): 
            if cell.value in colunas_formato:
                col_indices[cell.value] = c_idx
        if not col_indices:
            print("Aviso: Nenhuma coluna de formato encontrada no cabeçalho.")
            return workbook
        for row in ws.iter_rows(min_row=2):
            for col_name, col_idx in col_indices.items():
                cell = row[col_idx - 1] 
                if cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = colunas_formato[col_name]
                    except (ValueError, TypeError):
                        pass 
        print("Formatação aplicada com sucesso ao Workbook.")
        return workbook
    except Exception as e:
        print(f"Aviso: Falha ao aplicar formatação OpenPyXL. Erro: {e}")
        return workbook


# --- FUNÇÃO PRINCIPAL (MODIFICADA PARA LER DO FIREBASE) ---

def rodar_conciliacao_streamlit(caminho_arquivo_xml):
    """
    Executa a lógica de conciliação.
    AGORA LÊ a base de compras do FIREBASE.
    """

    # --- NOVO: Download do Arquivo B (do Firebase) ---
    print("Iniciando a conciliação...")
    print(f"Conectando ao Firebase para baixar a '{COLECAO_FIRESTORE}'...")
    db = get_db()
    if db is None:
        raise Exception("Não foi possível conectar ao Firestore.")
    
    docs_stream = db.collection(COLECAO_FIRESTORE).stream()
    dados_compras = [doc.to_dict() for doc in docs_stream]
    
    if not dados_compras:
        raise Exception(f"Nenhum dado encontrado em '{COLECAO_FIRESTORE}'. Você já carregou a Base de Compras (Passo 1)?")
    
    df_compras = pd.DataFrame(dados_compras)
    print(f"{len(df_compras)} registros da Base de Compras baixados do Firebase.")

    # --- 1. Leitura e Limpeza do Arquivo A (XML) ---
    # (Esta parte é idêntica à sua lógica anterior)
    print(f"Lendo '{caminho_arquivo_xml}'...")
    try:
        df_xml = read_spreadsheetml(caminho_arquivo_xml, 
                                  sheet_name="2-Titulos a pagar", 
                                  header_row=1)
        if df_xml.empty:
            raise Exception("O parser 'read_spreadsheetml' retornou um DataFrame vazio.")
        df_xml.columns = [str(c).replace('\n', ' ').strip() for c in df_xml.columns]
        df_xml['_xml_row_id'] = range(len(df_xml))
    except Exception as e:
        print(f"Erro ao ler o XML com 'read_spreadsheetml': {e}")
        raise e

    print("Arquivo XML lido. Tratando colunas-chave...")
    df_xml[['Código', 'Loja', 'Nome do Fornecedor']] = \
        df_xml['Codigo-Nome do Fornecedor'].apply(tratar_fornecedor)
    df_xml[['Parcela', 'Documento']] = \
        df_xml['Prf-Numero Parcela'].apply(tratar_prf_parcela)
    df_xml.dropna(subset=['Documento', 'Parcela'], inplace=True)
    colunas_texto_xml = ['Código', 'Documento', 'Parcela', 'Loja', 'Centro Custo', 'Cta.Contabil', 'Negocio?']
    for col in colunas_texto_xml:
        if col in df_xml.columns:
            df_xml[col] = df_xml[col].astype(str).str.strip().fillna('')
    
    colunas_finais_xml = list(df_xml.columns) + ['Vlr Rateado', 'Filial']
    xml_keys = ['Código', 'Documento'] 

    # --- 2. Preparação do Arquivo B (DataFrame do Firebase) ---
    # (Esta parte é idêntica à sua lógica anterior)
    
    col_forn_db = COLUNAS_CHAVE_EXCEL['codigo_fornecedor']
    col_doc_db = COLUNAS_CHAVE_EXCEL['numero_documento']

    if col_forn_db not in df_compras.columns or col_doc_db not in df_compras.columns:
        raise ValueError(f"DataFrame 'df_compras' do Firebase não contém colunas-chave: {col_forn_db}, {col_doc_db}")

    df_compras[col_forn_db] = df_compras[col_forn_db].astype(str).str.lstrip('0').str.strip().fillna('')
    df_compras[col_doc_db] = df_compras[col_doc_db].astype(str).str.lstrip('0').str.strip().fillna('')
    df_xml['Código'] = df_xml['Código'].astype(str).str.lstrip('0').str.strip().fillna('')
    df_xml['Documento'] = df_xml['Documento'].astype(str).str.lstrip('0').str.strip().fillna('')
    
    colunas_texto_db = ['Centro Custo', 'C Contabil', 'Item Conta', 'Loja', 'Filial']
    for col in colunas_texto_db:
        if col in df_compras.columns:
            df_compras[col] = df_compras[col].astype(str).str.strip().fillna('')
    
    # Garante que Vlr.Total é numérico (o Firebase deve ter mantido, mas por via das dúvidas)
    if 'Vlr.Total' in df_compras.columns:
        df_compras['Vlr.Total'] = pd.to_numeric(df_compras['Vlr.Total'], errors='coerce').fillna(0.0)
    else:
        print("Aviso: Coluna 'Vlr.Total' não encontrada no df_compras do Firebase.")
        df_compras['Vlr.Total'] = 0.0
    
    print("Pré-calculando contagem de rateios...")
    db_counts = df_compras.groupby([col_forn_db, col_doc_db]).size().to_frame('db_match_count')
    print("Pré-calculando soma total (custo) do rateio...")
    db_soma_doc = df_compras.groupby([col_forn_db, col_doc_db])['Vlr.Total'].sum().to_frame('Soma_Doc')
    df_compras = df_compras.merge(db_counts, left_on=[col_forn_db, col_doc_db], right_index=True, how='left')
    df_compras = df_compras.merge(db_soma_doc, left_on=[col_forn_db, col_doc_db], right_index=True, how='left')
        
    print(f"{len(df_compras)} registros da base de compras prontos.")

    # --- 3. A Lógica Central (Merge) ---
    # (Esta parte é idêntica à sua lógica anterior)
    print("Iniciando a conciliação (merge)...")
    df_merged = pd.merge(
        df_xml,
        df_compras,
        left_on=xml_keys,
        right_on=[col_forn_db, col_doc_db],
        how='left', 
        indicator=True,
        suffixes=('_xml', '_db') 
    )
    df_merged['db_match_count'] = df_merged['db_match_count'].fillna(0)
    df_merged['Soma_Doc'] = df_merged['Soma_Doc'].fillna(0)

    # --- 4. Separação (Lógica idêntica) ---
    print("Separando linhas com e sem rateio...")
    def renomear_colunas(df):
        cols_para_renomear = {col: col.replace('_xml', '') for col in df.columns if col.endswith('_xml')}
        df.rename(columns=cols_para_renomear, inplace=True)
        return df

    condicao_sem_rateio = (df_merged['db_match_count'] <= 1)
    df_final_sem_rateio = df_merged[condicao_sem_rateio].copy()
    df_final_sem_rateio.drop_duplicates(subset=['_xml_row_id'], keep='first', inplace=True)
    df_final_sem_rateio = renomear_colunas(df_final_sem_rateio)
    df_final_sem_rateio['Vlr Rateado'] = df_final_sem_rateio['Titulos a vencer Valor nominal'].apply(to_number_brl)
    df_final_sem_rateio = df_final_sem_rateio[colunas_finais_xml]

    condicao_com_rateio = (df_merged['db_match_count'] > 1)
    df_final_com_rateio = df_merged[condicao_com_rateio].copy()

    # --- 5. Mapeamento e AGRUPAMENTO (Lógica idêntica, incluindo "Filial") ---
    if not df_final_com_rateio.empty:
        print("Agrupando e somando títulos rateados...")
        col_titulos_vencer_xml = 'Titulos a vencer Valor nominal_xml' if 'Titulos a vencer Valor nominal_xml' in df_final_com_rateio.columns else 'Titulos a vencer Valor nominal'
        df_final_com_rateio['Valor_Pago_Num'] = df_final_com_rateio[col_titulos_vencer_xml].apply(to_number_brl)
        
        col_cc_db = 'Centro Custo_db' if 'Centro Custo_db' in df_final_com_rateio.columns else 'Centro Custo'
        col_cta_db = 'C Contabil' 
        col_item_db = 'Item Conta' 
        col_loja_db = 'Loja_db' if 'Loja_db' in df_final_com_rateio.columns else 'Loja'
        col_vlr_db = 'Vlr.Total' 
        col_filial_db = 'Filial' 
        grouping_keys = ['_xml_row_id', col_item_db, col_cc_db, col_cta_db, col_loja_db, col_filial_db]
        
        agg_funcs = { 
            col_vlr_db: 'sum', 
            'Soma_Doc': 'first', 
            'Valor_Pago_Num': 'first',
            'Código': 'first', 
            'Documento': 'first',
            'Parcela': 'first',
            'Filial': 'first'
        }
        
        xml_cols_to_keep = [col for col in colunas_finais_xml if col not in ['Código', 'Documento', 'Parcela', 'Vlr Rateado', '_xml_row_id', 'Filial']]
        for col in df_final_com_rateio.columns:
            if col.endswith('_xml') and col not in agg_funcs: agg_funcs[col] = 'first'
            elif col in xml_cols_to_keep and col not in agg_funcs: agg_funcs[col] = 'first'

        grouping_keys_existentes = [key for key in grouping_keys if key in df_final_com_rateio.columns]
        
        print(f"Agrupando rateios por {grouping_keys_existentes}...")
        df_agrupado = df_final_com_rateio.groupby(grouping_keys_existentes, as_index=False).agg(agg_funcs)

        df_agrupado['Proporcao'] = 0.0
        mask_soma_valida = df_agrupado['Soma_Doc'] != 0
        df_agrupado.loc[mask_soma_valida, 'Proporcao'] = df_agrupado[col_vlr_db] / df_agrupado['Soma_Doc']
        df_agrupado['Vlr Rateado'] = df_agrupado['Proporcao'] * df_agrupado['Valor_Pago_Num']
        df_agrupado['Valor Original'] = df_agrupado[col_vlr_db]
        
        if col_cc_db in df_agrupado.columns: df_agrupado['Centro Custo_xml'] = df_agrupado[col_cc_db]
        if col_cta_db in df_agrupado.columns:
            col_cta_xml = 'Cta.Contabil_xml' if 'Cta.Contabil_xml' in df_agrupado.columns else 'Cta.Contabil'
            df_agrupado[col_cta_xml] = df_agrupado[col_cta_db]
        if col_item_db in df_agrupado.columns:
            col_negocio_xml = 'Negocio?_xml' if 'Negocio?_xml' in df_agrupado.columns else 'Negocio?'
            df_agrupado[col_negocio_xml] = df_agrupado[col_item_db]
        if col_loja_db in df_agrupado.columns: df_agrupado['Loja_xml'] = df_agrupado[col_loja_db]
        
        df_final_com_rateio = renomear_colunas(df_agrupado)
        df_final_com_rateio = df_final_com_rateio[colunas_finais_xml]
    else:
        print("Nenhum título com rateio (múltiplas linhas) foi encontrado na base.")

    # --- 6. Junção e Finalização (Lógica idêntica) ---
    print("Consolidando relatório final...")
    df_final = pd.concat([df_final_sem_rateio, df_final_com_rateio], ignore_index=True)
    
    print("Formatando colunas de data para DD/MM/YYYY...")
    colunas_data = ['Data de Emissao', 'Data de Vencto', 'Vencto Real']
    for col in colunas_data:
        if col in df_final.columns:
            df_final[col] = df_final[col].apply(formatar_data_br)
    
    colunas_novas = ['Código', 'Loja', 'Nome do Fornecedor', 'Documento', 'Parcela']
    colunas_originais_xml = list(df_xml.drop(columns=colunas_novas + ['_xml_row_id'], errors='ignore').columns)
    
    ordem_final = colunas_novas + colunas_originais_xml + ['Vlr Rateado', 'Filial']
    colunas_existentes_na_ordem = [col for col in ordem_final if col in df_final.columns]
    
    df_final = df_final[colunas_existentes_na_ordem]
    
    print(f"Salvando relatório em memória (BytesIO)...")
    
    output_stream = io.BytesIO()
    df_final.to_excel(output_stream, index=False, engine='openpyxl')
    output_stream.seek(0) 
    
    colunas_para_formatar = {
        'Valor Original': ACCOUNTING_FORMAT,
        'Tit Vencidos Valor nominal': ACCOUNTING_FORMAT,
        'Titulos a vencer Valor nominal': ACCOUNTING_FORMAT,
        'Vlr Rateado': ACCOUNTING_FORMAT 
    }
    
    if OPENPYXL_DISPONIVEL:
        try:
            wb = load_workbook(output_stream)
            wb_formatado = aplicar_formatacao_excel(wb, colunas_para_formatar)
            output_stream_formatado = io.BytesIO()
            wb_formatado.save(output_stream_formatado)
            output_stream_formatado.seek(0)
            
            print("\n--- SUCESSO! ---")
            print("Relatório final gerado e formatado em memória.")
            return output_stream_formatado, True 
            
        except Exception as e:
            print(f"Falha ao formatar o Excel em memória, retornando não formatado. Erro: {e}")
            output_stream.seek(0)
            return output_stream, False 
    else:
        output_stream.seek(0)
        return output_stream, False



